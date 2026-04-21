"""

SR Report Generator - Product Ops



File size gate:

  > 60 MB  →  Report Generator only  (Overview / Daily / Weekly / Monthly)

  ≤ 60 MB  →  Report Generator + Insights & RCA + Smart Summary



Performance features:

  - Polars implementation for lighting-fast I/O and vectorization.

  - Native Polars Lazy-style aggregation to prevent RAM exhaustion.

  - Single-pass groupby in compute_sr.

  - Pandas conversion only applied at the Excel/UI boundary (tiny datasets).

"""



import io

import gc

from concurrent.futures import ThreadPoolExecutor, as_completed

from typing import Optional, Dict, List



import polars as pl

import pandas as pd

import streamlit as st

import plotly.graph_objects as go

from openpyxl import load_workbook

from openpyxl.styles import Font

from openpyxl.formatting.rule import FormulaRule, DataBarRule

from openpyxl.utils import get_column_letter



# ---------------------------------------------------------------------------

# PAGE CONFIG

# ---------------------------------------------------------------------------

st.set_page_config(page_title="Payment SR Analysis", layout="wide")

st.title("SR Report Generator - Product Ops")



st.markdown("""

<a href="https://metabase.cashfree.com/question/23625-sr-report-data?merchantid=&start_date=&end_date="

   target="_blank" style="font-size:18px;font-weight:bold;">

CLICK HERE TO GET THE DATA FROM METABASE

</a>

""", unsafe_allow_html=True)

st.markdown("---")



# ---------------------------------------------------------------------------

# CONSTANTS

# ---------------------------------------------------------------------------

LARGE_FILE_THRESHOLD_MB = 60   # files above this get Report Generator only



TIER_1_BANKS = frozenset([

    "AXIS BANK", "HDFC BANK", "ICICI BANK",

    "KOTAK MAHINDRA BANK", "STATE BANK OF INDIA", "YES BANK LTD",

])

CARD_MODES = frozenset([

    "CREDIT_CARD", "DEBIT_CARD", "CARD", "PREPAID_CARD", "CREDIT_CARD_EMI",

])



PSP_MAP = {

    "abcdicici": "ABCD (Aditya Birla Capital)", "abfspay": "Bajaj Finserv / Markets",

    "airtel": "Airtel Thanks App", "allbank": "Allahabad Bank (Now Indian Bank)",

    "amazonpay": "Amazon Pay", "andb": "Andhra Bank (Now Union Bank)", "apl": "Amazon Pay",

    "ary": "Aryavart Bank", "aubank": "AU Small Finance Bank", "axisb": "Axis Mobile / CRED",

    "axisbank": "Axis Mobile App", "axl": "PhonePe", "bandhan": "Bandhan Bank",

    "barodampay": "bob World", "barodapay": "bob World", "bhim": "BHIM App",

    "boi": "BHIM Aadhaar / Mobile", "bombob": "BHIM BOM", "bpaywallet": "BharatPe",

    "bpunity": "BharatPe", "cbin": "Central Bank Mobile", "centralbank": "Central Bank Mobile",

    "citi": "Citibank India (Now Axis)", "citigold": "Citibank India (Now Axis)",

    "cnrb": "Candy / Canara ai1", "corp": "Corporation Bank (Now Union Bank)", "cred": "CRED",

    "csbpay": "CSB Bank", "cub": "CUB mBank Plus", "db": "Deutsche Bank",

    "dbs": "DBS Digibank", "dhani": "Dhani App", "digikhata": "DigiKhata",

    "dlb": "Dhanlaxmi Bank", "dnsbank": "DNS Bank", "draxisbank": "DR", "drbob": "DR",

    "drcanb": "DR", "drfederal": "DR", "drhdfcbank": "DR", "dricici": "DR",

    "dridbi": "DR", "dridfc": "DR", "drindus": "DR", "drkotak": "DR", "drpnb": "DR",

    "drsbi": "DR", "drubi": "DR", "druco": "DR", "dryesb": "DR", "ebixcash": "EbixCash",

    "equitas": "Equitas Mobile", "equitasbank": "Equitas Mobile", "esaf": "ESAF Mobile",

    "fam": "FamPay", "fbl": "FedMobile / CoinTab", "federal": "FedMobile",

    "fifederal": "Fi Money", "fincarebank": "Fincare Mobile", "finobank": "FinoPay",

    "fkaxis": "Flipkart", "freecharge": "Freecharge", "freoicici": "Freo",

    "goaxb": "Kiwi", "gwaxis": "Genwise", "hdfc": "HDFC Bank Mobile",

    "hdfcbank": "HDFC Bank Mobile", "hsbc": "HSBC India", "hsbcbank": "HSBC India",

    "ibl": "PhonePe", "icici": "iMobile / Pockets", "idbi": "IDBI Go Mobile+",

    "idfcbank": "IDFC First Mobile", "idfcfirst": "IDFC First Mobile", "ikwik": "MobiKwik",

    "imobile": "iMobile Pay", "indianbank": "IndOASIS", "indianbk": "IndOASIS",

    "indie": "INDIE", "indus": "IndusMobile", "indusind": "IndusMobile",

    "inhdfc": "Tata Neu", "iob": "IOB Mobile", "janabank": "Jana Mobile",

    "jarunity": "Jar App", "jio": "MyJio / JioPay", "jkb": "J&K Bank",

    "jsb": "Janaseva Bank", "jupiter": "Jupiter Money", "jupiteraxis": "Jupiter Money",

    "kaypay": "Kotak (Legacy)", "kbaxis": "KreditBee", "kbl": "KBL Mobile",

    "kotak": "Kotak Mobile", "kotak811": "Kotak 811", "kphdfc": "Kredit.Pe",

    "kvb": "KVB Dlite", "liv": "LivQuik", "lxaxis": "LiquiLoans", "mahb": "MahaMobile",

    "maxaxis": "Max Life", "mbk": "MobiKwik", "mbkns": "MobiKwik",

    "mboi": "Bank of India Mobile", "mvhdfc": "Money View", "naviaxis": "Navi App",

    "niyoicici": "Niyo", "nsdl": "NSDL Jiffy", "nye": "Niyo", "nyes": "Niyo",

    "obopay": "Obopay", "okaxis": "Google Pay", "okhdfcbank": "Google Pay",

    "okicici": "Google Pay", "oksbi": "Google Pay", "omni": "OmniCard",

    "oneyes": "OneCard", "paytm": "Paytm", "paytmwallet": "Paytm", "payu": "PayU",

    "payworld": "Payworld", "payzapp": "PayZapp", "phonepe": "PhonePe",

    "pinelabs": "Pine Labs", "pingpay": "Samsung Pay Mini", "pnb": "PNB One",

    "pnyes": "PennyDrop", "pockets": "Pockets", "postbank": "IPPB Mobile",

    "psb": "PSB UnIC", "psbank": "PSB UnIC", "ptaxis": "Paytm", "pthdfc": "Paytm",

    "ptsbi": "Paytm", "ptyes": "Paytm", "pz": "PayZapp", "pzh": "PayZapp",

    "pzw": "PayZapp", "rapl": "Amazon Pay", "razorpay": "Razorpay",

    "rbl": "RBL MoBank", "rmrbl": "Resilient", "sbi": "BHIM SBI Pay / Yono",

    "sbmbank": "SBM Bank", "scb": "SC Mobile", "seyes": "SalarySe",

    "shriramhdfcbank": "Shriram Finance", "sib": "Mirror+", "slc": "Slice",

    "slice": "Slice", "sliceaxis": "Slice", "slicepay": "Slice",

    "spicepay": "Spice Money", "superyes": "Super.Money",

    "suryoday": "Suryoday Mobile", "tapicici": "Tata Neu",

    "tbl": "Thane Bharat Bank", "timecosmos": "TimePay", "tjsb": "TJSB Mobile",

    "tmb": "TMB Digilobby", "topay": "ToPay", "trans": "Cheq / Transcorp",

    "trio": "Trio", "ubi": "Union Bank", "uboi": "Union Bank", "uco": "UCO mBanking",

    "ujjivan": "Ujjivan Mobile", "unionbank": "Union Bank",

    "unionbankofindia": "Union Bank", "unitypay": "Unity Bank", "upi": "BHIM",

    "utkarshbank": "Utkarsh Mobile", "waaxis": "WhatsApp Pay",

    "wahdfcbank": "WhatsApp Pay", "waicici": "WhatsApp Pay", "wasbi": "WhatsApp Pay",

    "yapl": "Amazon Pay", "ybl": "PhonePe", "yes": "Yes Bank",

    "yesbank": "Iris by Yes Bank", "yescred": "CRED", "yescurie": "CRED (Curie)",

    "yesfam": "FamPay", "yesg": "Groww Pay", "yesgo": "Yes Bank",

    "yespay": "Yes Pay Next", "yespop": "POP", "yestp": "Third Party App",

    "zoicici": "Zomato", "ztrbl": "Zeta",

}



# ---------------------------------------------------------------------------

# LOAD + PREPROCESS — Polars Engine

# ---------------------------------------------------------------------------

@st.cache_data(show_spinner=False)

def load_and_preprocess(file_bytes: bytes) -> pl.DataFrame:

    # 1. Load via Polars - Force all columns to Strings initially to prevent schema crash

    df = pl.read_csv(

        file_bytes,

        infer_schema_length=0,  # CRITICAL: Reads everything as text first like Pandas

        null_values=["NAN", "NONE", "nan", "None", "", "NaN"],

        ignore_errors=True,

        truncate_ragged_lines=True # Protects against messy CSV rows

    )



    # 2. Clean Column Names

    df = df.rename({c: c.strip().lower().replace(" ", "_") for c in df.columns})



    # Clean merchantid commas (Metabase sometimes exports 6,04,088 instead of 604088)

    if "merchantid" in df.columns:

        df = df.with_columns(pl.col("merchantid").cast(pl.Utf8).str.replace_all(",", ""))



    # 3. Clean categorical text

    cat_cols = ["paymentmode", "txstatus", "bankname", "cardtype", "cardcountry", "pg", "txmsg"]

    exprs = []

    for c in cat_cols:

        if c in df.columns:

            exprs.append(

                pl.col(c).cast(pl.Utf8).str.strip_chars().str.to_uppercase()

                .replace({"NAN": None, "NONE": None})

                .cast(pl.Categorical)

            )

    df = df.with_columns(exprs)



    # 4. Robust Datetime parsing

    if "txtime" in df.columns:

        # Normalize the string: remove commas, T separators, and extra spaces

        df = df.with_columns(

            pl.col("txtime").cast(pl.Utf8)

            .str.replace_all("T", " ")

            .str.replace_all(",", "")

            .str.replace_all(r"\s+", " ")

            .str.strip_chars()

        )

        

        # Try multiple formats sequentially. Strict=False returns nulls instead of crashing.

        df = df.with_columns(

            pl.coalesce([

                # NEW METABASE FORMATS (e.g., "January 1 2026 12:41 PM" or "January 1 2026 4:44 AM")

                pl.col("txtime").str.to_datetime("%B %-d %Y %-I:%M %p", strict=False),

                pl.col("txtime").str.to_datetime("%B %d %Y %I:%M %p", strict=False),

                pl.col("txtime").str.to_datetime("%B %-d %Y %I:%M %p", strict=False),

                pl.col("txtime").str.to_datetime("%B %d %Y %-I:%M %p", strict=False),

                

                # STANDARD FORMATS

                pl.col("txtime").str.to_datetime("%Y-%m-%d %H:%M:%S", strict=False),

                pl.col("txtime").str.to_datetime("%Y-%m-%d %H:%M:%S.%f", strict=False),

                pl.col("txtime").str.to_datetime("%d-%m-%Y %H:%M:%S", strict=False),

                pl.col("txtime").str.to_datetime("%d/%m/%Y %H:%M:%S", strict=False),

                pl.col("txtime").str.to_datetime("%Y-%m-%d %H:%M", strict=False),

                pl.col("txtime").str.to_datetime("%Y/%m/%d %H:%M:%S", strict=False),

                pl.col("txtime").str.to_datetime("%d-%m-%Y %H:%M", strict=False),

            ]).alias("txtime")

        ).drop_nulls(subset=["txtime"])



        # Create derived time columns

        df = df.with_columns([

            pl.col("txtime").dt.date().alias("Day"),

            pl.col("txtime").dt.strftime("%Y-%m").cast(pl.Categorical).alias("Month"),

            # FIXED: Week column now shows date range (e.g., "01-01-2026 to 07-01-2026")

            (

                pl.col("txtime").dt.offset_by("-" + pl.col("txtime").dt.weekday().cast(pl.String) + "d")

                .dt.date()

                .dt.strftime("%d-%m-%Y")

                + " to " +

                pl.col("txtime").dt.offset_by((6 - pl.col("txtime").dt.weekday()).cast(pl.String) + "d")

                .dt.date()

                .dt.strftime("%d-%m-%Y")

            ).cast(pl.Categorical).alias("Week"),

            pl.col("txtime").dt.strftime("%Y-%m-%d %H:00:00").cast(pl.Categorical).alias("Hour"),

            pl.col("txtime").dt.strftime("%d-%b").cast(pl.Categorical).alias("Display_Date"),

        ]).drop("txtime")



    # 5. Amount parsing

    if "amount" in df.columns:

        df = df.with_columns(

            pl.col("amount").cast(pl.Utf8).str.replace_all(r"[^\d.]", "")

            .cast(pl.Float32, strict=False).fill_null(0.0)

        )

    else:

        df = df.with_columns(pl.lit(0.0, dtype=pl.Float32).alias("amount"))



    df = df.with_columns(

        pl.when(pl.col("amount") <= 1000).then(pl.lit("0-1k"))

          .when(pl.col("amount") <= 10000).then(pl.lit("1k-10k"))

          .when(pl.col("amount") <= 50000).then(pl.lit("10k-50k"))

          .when(pl.col("amount") <= 100000).then(pl.lit("50k-1L"))

          .when(pl.col("amount") <= 200000).then(pl.lit("1L-2L"))

          .otherwise(pl.lit(">2L"))

          .cast(pl.Categorical)

          .alias("amount_category")

    )



    # 6. Boolean checks

    df = df.with_columns([

        (pl.col("txstatus") == "SUCCESS").cast(pl.Int8).alias("is_success"),

        (pl.col("txstatus") == "USER_DROPPED").cast(pl.Int8).alias("is_userdrop"),

        (pl.col("txstatus") == "FAILED").cast(pl.Int8).alias("is_failed"),

        (pl.col("txstatus").is_in(["PENDING", "INCOMPLETE", "FLAGGED", "CANCELLED"])).cast(pl.Int8).alias("is_incomplete"),

    ])



    # 7. Bank tier

    if "bankname" in df.columns:

        df = df.with_columns(

            pl.when(pl.col("bankname").is_in(list(TIER_1_BANKS)))

              .then(pl.lit("Tier 1 Bank"))

              .otherwise(pl.lit("Tier 2 Bank"))

              .cast(pl.Categorical)

              .alias("bank_tier")

        )



    # 8. UPI & Card mappings (Regex extraction prevents Out-Of-Bounds index errors)

    if "cardnumber" in df.columns:

        df = df.with_columns(

            pl.col("cardnumber").cast(pl.Utf8)

              .str.extract(r"@(.*)", 1)

              .str.to_lowercase()

              .alias("upi_handle")

        )

        df = df.with_columns(

            pl.col("upi_handle").replace(PSP_MAP, default=pl.col("upi_handle")).cast(pl.Categorical).alias("psp_app"),

            pl.col("upi_handle").cast(pl.Categorical)

        ).drop("cardnumber")



    if "cardcountry" in df.columns:

        df = df.with_columns(

            pl.when(pl.col("cardcountry") == "IN").then(pl.lit("DOMESTIC")).otherwise(pl.lit("IPG"))

            .cast(pl.Categorical).alias("card_category")

        )



    # 9. Sub-Category Vectorization

    pm = pl.col("paymentmode") if "paymentmode" in df.columns else pl.lit("")

    bn = pl.col("bankname") if "bankname" in df.columns else pl.lit(None)

    cc = pl.col("cardcountry") if "cardcountry" in df.columns else pl.lit("IN")



    has_bank = bn.is_not_null() & ~bn.is_in(["NAN", "NONE", ""])

    is_dom = cc == "IN"



    sub_cat_expr = pl.when(pm == "UPI").then(

        pl.when(has_bank).then(pl.lit("UPI_INTENT")).otherwise(pl.lit("UPI_COLLECT"))

    )

    for mode in CARD_MODES:

        sub_cat_expr = sub_cat_expr.when((pm == mode) & is_dom).then(pl.lit(f"{mode}_DOMESTIC"))

        sub_cat_expr = sub_cat_expr.when((pm == mode) & ~is_dom).then(pl.lit(f"{mode}_INTERNATIONAL"))



    sub_cat_expr = sub_cat_expr.when(pm == "NET_BANKING").then(

        pl.when(bn.is_in(list(TIER_1_BANKS))).then(pl.lit("NB_TIER_1")).otherwise(pl.lit("NB_TIER_2"))

    )

    sub_cat_expr = sub_cat_expr.otherwise(pl.lit("OTHER")).cast(pl.Categorical)

    df = df.with_columns(sub_cat_expr.alias("sub_category"))



    # 10. Pre-compute Groupby metrics

    df = df.with_columns([

        (pl.col("amount") * pl.col("is_success")).cast(pl.Float32).alias("success_amount"),

        (pl.col("txstatus") != "USER_DROPPED").cast(pl.Int8).alias("nodrop_flag")

    ])

    df = df.with_columns(

        (pl.col("is_success") * pl.col("nodrop_flag")).cast(pl.Int8).alias("nodrop_success")

    )



    return df





# ---------------------------------------------------------------------------

# COMPUTE SR — Polars Engine

# ---------------------------------------------------------------------------

_AGG_COLS = ["Volume","Success","UserDrops","Total_Value","GMV","nd_succ","nd_vol"]



def _build_base(data: pl.DataFrame, key: list, tx_id_col: str, amount_col: str) -> pl.DataFrame:

    return data.group_by(key, maintain_order=False).agg([

        pl.len().alias("Volume"), # count of rows

        pl.col("is_success").sum().alias("Success"),

        pl.col("is_userdrop").sum().alias("UserDrops"),

        pl.col(amount_col).sum().alias("Total_Value"),

        pl.col("success_amount").sum().alias("GMV"),

        pl.col("nodrop_success").sum().alias("nd_succ"),

        pl.col("nodrop_flag").sum().alias("nd_vol"),

    ])



def _rollup(base: pl.DataFrame, key: list) -> pl.DataFrame:

    return base.group_by(key, maintain_order=False).agg([

        pl.col(c).sum().alias(c) for c in _AGG_COLS

    ])



def _finalise(grouped: pl.DataFrame, total_volume: int, merchant_col: str, 

              num_merchants: int, merchant_totals: Optional[pl.DataFrame]) -> pl.DataFrame:

    grouped = grouped.with_columns([

        (pl.col("Volume") - pl.col("Success")).alias("Unsuccessful Count"),

        (pl.col("Success") / pl.when(pl.col("Volume")==0).then(1).otherwise(pl.col("Volume")) * 100).round(2).alias("SR (%)"),

        (pl.col("nd_succ") / pl.when(pl.col("nd_vol")==0).then(1).otherwise(pl.col("nd_vol")) * 100).round(2).alias("SR without User Drops (%)"),

        (pl.col("Volume") / total_volume * 100).round(2).alias("% of Volume (Global)")

    ]).drop(["nd_succ", "nd_vol"])



    if num_merchants > 1 and merchant_totals is not None:

        grouped = grouped.join(merchant_totals, on=merchant_col, how="left")

        grouped = grouped.with_columns(

            (pl.col("Volume") / pl.when(pl.col("_mt")==0).then(1).otherwise(pl.col("_mt")) * 100).round(2).alias("% of Volume (Per Merchant)")

        ).drop("_mt")



    return grouped.sort("Volume", descending=True)



def compute_sr(data: pl.DataFrame, group_cols, merchant_col,

               tx_id_col, amount_col, status_col, num_merchants,

               _base_cache: Optional[dict] = None) -> pl.DataFrame:

    group_cols = [c for c in group_cols if c is not None]

    total_volume = data.height

    if total_volume == 0:

        return pl.DataFrame()



    key = [merchant_col] + group_cols



    if _base_cache is not None:

        cache_key = id(data)

        if cache_key not in _base_cache:

            _base_cache[cache_key] = _build_base(data, key, tx_id_col, amount_col)

        base = _base_cache[cache_key]

        base_key_set = set(base.columns) - set(_AGG_COLS)

        if set(key).issubset(base_key_set):

            grouped = _rollup(base, key)

        else:

            grouped = _build_base(data, key, tx_id_col, amount_col)

    else:

        grouped = _build_base(data, key, tx_id_col, amount_col)



    mt = None

    if num_merchants > 1:

        mt = data.group_by(merchant_col).agg(pl.len().alias("_mt"))



    return _finalise(grouped, total_volume, merchant_col, num_merchants, mt)



def compute_mom_change(df: pl.DataFrame, group_cols, merchant_col) -> pl.DataFrame:

    if df.is_empty(): return df

    group_cols = [c for c in group_cols if c is not None]

    key_cols = [c for c in group_cols if c not in ("Month", "Week", "Hour")]



    df = df.sort([merchant_col] + group_cols)

    part_cols = [merchant_col] + key_cols



    exprs = []

    if "Volume" in df.columns:

        exprs.extend([

            pl.col("Volume").diff().over(part_cols).fill_null(0).alias("Volume Δ"),

            (pl.col("Volume").pct_change().over(part_cols).fill_null(0) * 100).round(2).alias("Volume % Change")

        ])

    if "SR (%)" in df.columns:

        exprs.extend([

            pl.col("SR (%)").diff().over(part_cols).fill_null(0).alias("SR (%) Δ"),

            (pl.col("SR (%)").pct_change().over(part_cols).fill_null(0) * 100).round(2).alias("SR (%) % Change")

        ])



    if exprs:

        df = df.with_columns(exprs)

    return df





# ---------------------------------------------------------------------------

# EXCEL FORMATTING

# ---------------------------------------------------------------------------

def _format_workbook(wb) -> io.BytesIO:

    for ws in wb.worksheets:

        if ws.max_row <= 1:

            continue

        headers = [c.value for c in ws[1]]

        vol_idx = headers.index("Volume") if "Volume" in headers else None

        vol_ltr = get_column_letter(vol_idx + 1) if vol_idx is not None else None



        for ci, hdr in enumerate(headers, 1):

            cl  = get_column_letter(ci)

            rng = f"{cl}2:{cl}{ws.max_row}"

            if hdr in ("SR (%)", "SR without User Drops (%)") and vol_ltr:

                ws.conditional_formatting.add(rng, FormulaRule(

                    formula=[f"AND({vol_ltr}2>=10,{cl}2<50)"], font=Font(color="FF0000")))

                ws.conditional_formatting.add(rng, FormulaRule(

                    formula=[f"AND({vol_ltr}2>=10,{cl}2>90)"], font=Font(color="008000")))

            elif hdr == "Unsuccessful Count" and vol_ltr:

                ws.conditional_formatting.add(rng, FormulaRule(

                    formula=[f"AND({vol_ltr}2>=10,{cl}2>=20,{cl}2>0.5*{vol_ltr}2)"],

                    font=Font(color="FF4500")))

            elif isinstance(hdr, str) and "Volume" in hdr and "%" in hdr and "Change" not in hdr:

                ws.conditional_formatting.add(rng, DataBarRule(

                    start_type="num", start_value=0,

                    end_type="num",   end_value=100, color="638EC6"))



        for col in ws.iter_cols():

            ltr = col[0].column_letter

            mx  = max((len(str(c.value or "")) for c in col[:100]), default=8)

            ws.column_dimensions[ltr].width = min(mx + 2, 50)



    out = io.BytesIO()

    wb.save(out)

    out.seek(0)

    return out



def apply_formatting(buffer: io.BytesIO) -> io.BytesIO:

    buffer.seek(0)

    wb = load_workbook(buffer)

    return _format_workbook(wb)





# ---------------------------------------------------------------------------

# BUILD EXCEL REPORT FROM POLARS

# ---------------------------------------------------------------------------

def _build_report_excel(

    sheet_specs, base_breakdowns, cur: pl.DataFrame, tcol, report_type,

    num_merchants, hourly_df=None

) -> io.BytesIO:

    

    FINEST_COLS = ["merchantid","paymentmode","pg","bankname",

                   "sub_category","upi_handle","psp_app","bank_tier",

                   "card_category","cardtype","amount_category"]

    if tcol:

        FINEST_COLS.append(tcol)



    dataset_bases: dict = {}

    for sn, (cols, ds) in sheet_specs.items():

        did = id(ds)

        if did not in dataset_bases:

            finest_key = [c for c in FINEST_COLS if c in ds.columns]

            if not finest_key:

                finest_key = ["merchantid"]

            dataset_bases[did] = _build_base(ds, finest_key, "transactionid", "amount")



    def _sheet_from_base(sn, cols, ds):

        did = id(ds)

        base = dataset_bases[did]

        key = ["merchantid"] + [c for c in cols if c is not None]

        total = int(base.select(pl.col("Volume").sum()).item())

        if total == 0: return sn, None



        if all(c in base.columns for c in key):

            grouped = _rollup(base, key)

        else:

            grouped = _build_base(ds, key, "transactionid", "amount")



        mt = None

        if num_merchants > 1:

            mt = base.group_by("merchantid").agg(pl.col("Volume").sum().alias("_mt"))



        result = _finalise(grouped, total, "merchantid", num_merchants, mt)

        if result.is_empty(): return sn, None

        if tcol and tcol in result.columns:

            result = result.sort(tcol)

        if report_type in ("Monthly","Weekly"):

            result = compute_mom_change(result, cols, "merchantid")

        

        # Boundary crossover: Convert the TINY aggregated result to Pandas for Openpyxl

        return sn, result.to_pandas()



    computed: dict = {}

    

    # Polars multi-threads natively, so we run this sequentially to avoid thread contention.

    for sn, (cols, ds) in sheet_specs.items():

        sheet_name, res = _sheet_from_base(sn, cols, ds)

        if res is not None:

            computed[sheet_name] = res



    # ── failure sheets ────────────────

    fail_data = cur.filter(pl.col("txstatus") != "SUCCESS")

    if fail_data.height > 0:

        fc = ["merchantid"] + ([tcol] if tcol else []) + ["paymentmode","txmsg"]

        sort_c = [tcol] if tcol else ["Volume"]

        desc_s = [False] if tcol else [True]

        computed["Failures Analysis"] = fail_data.group_by(fc).agg(pl.len().alias("Volume")).sort(sort_c, descending=desc_s).to_pandas()



        if "bankname" in fail_data.columns:

            bc = ["merchantid"] + ([tcol] if tcol else []) + ["paymentmode","bankname","txmsg"]

            computed["Failures (Paymode+Bank)"] = fail_data.group_by(bc).agg(pl.len().alias("Volume")).sort("Volume", descending=True).to_pandas()



        if "pg" in fail_data.columns:

            pc = ["merchantid"] + ([tcol] if tcol else []) + ["paymentmode","pg","txmsg"]

            computed["Failures (Paymode+PG)"] = fail_data.group_by(pc).agg(pl.len().alias("Volume")).sort("Volume", descending=True).to_pandas()



    if not computed:

        computed["No Data"] = pd.DataFrame([{"Info": "No data"}])



    EXCEL_MAX_ROWS = 1_048_575

    split_log = []



    def _write_df_safe(writer, df_, sheet_name):

        total_rows = len(df_)

        if total_rows <= EXCEL_MAX_ROWS:

            df_.to_excel(writer, sheet_name=sheet_name[:31], index=False)

            return



        if "Volume" in df_.columns:

            df_ = df_.sort_values("Volume", ascending=False).reset_index(drop=True)



        total_parts = (total_rows + EXCEL_MAX_ROWS - 1) // EXCEL_MAX_ROWS

        base = sheet_name[:22]



        for part in range(total_parts):

            start = part * EXCEL_MAX_ROWS

            end   = min(start + EXCEL_MAX_ROWS, total_rows)

            chunk = df_.iloc[start:end]

            sn_part = f"{base} ({part+1} of {total_parts})"

            chunk.to_excel(writer, sheet_name=sn_part, index=False)



        split_log.append((sheet_name, total_parts, total_rows))



    out = io.BytesIO()

    with pd.ExcelWriter(out, engine="openpyxl") as writer:

        for sn, df_ in computed.items():

            _write_df_safe(writer, df_, sn[:31])



        if split_log:

            index_rows = []

            for base_name, n_parts, total_rows in split_log:

                index_rows.append({

                    "Sheet (base name)": base_name,

                    "Total Rows": total_rows,

                    "Split into N sheets": n_parts,

                    "Rows per sheet": EXCEL_MAX_ROWS,

                })

            pd.DataFrame(index_rows).to_excel(

                writer, sheet_name="⚠ Split Sheet Index", index=False)



        wb = writer.book

        _format_workbook(wb)



    out.seek(0)

    return out





# ---------------------------------------------------------------------------

# UI: REPORT TAB (POLARS COMPATIBLE)

# ---------------------------------------------------------------------------

def render_report_tab(df: pl.DataFrame):

    st.header("Generate Standard Excel Reports")

    c1, c2 = st.columns(2)

    with c1: share_with_merchant = st.checkbox("Share with Merchant (Remove PG Data)", value=False)

    with c2: run_hourly          = st.checkbox("Get Hourly SR Report")



    df_display = df.drop("pg") if share_with_merchant and "pg" in df.columns else df

    if share_with_merchant:

        st.warning("PG Data removed from reports.")



    hourly_df = None

    if run_hourly:

        mn = df_display.select(pl.col("Day").min()).item()

        mx = df_display.select(pl.col("Day").max()).item()

        hr = st.date_input("Select Date for Hourly Report", [mn, mx], min_value=mn, max_value=mx)

        if len(hr) == 2:

            hourly_df = df_display.filter((pl.col("Day") >= hr[0]) & (pl.col("Day") <= hr[1]))



    with st.expander("Filter Data (Date, Merchant, Mode)", expanded=True):

        if st.checkbox("Filter by Date Range", value=False):

            mn = df_display.select(pl.col("Day").min()).item()

            mx = df_display.select(pl.col("Day").max()).item()

            dr = st.date_input("Select Date Range", [mn, mx])

            if len(dr) == 2:

                df_display = df_display.filter((pl.col("Day") >= dr[0]) & (pl.col("Day") <= dr[1]))



        merchants = sorted(str(x) for x in df_display.get_column("merchantid").drop_nulls().unique().to_list())

        sel_merch = st.multiselect("Select Merchants", merchants, default=[])

        if sel_merch:

            df_display = df_display.filter(pl.col("merchantid").cast(pl.Utf8).is_in(sel_merch))

            if hourly_df is not None:

                hourly_df = hourly_df.filter(pl.col("merchantid").cast(pl.Utf8).is_in(sel_merch))



        if "paymentmode" in df_display.columns:

            modes     = sorted(str(x) for x in df_display.get_column("paymentmode").drop_nulls().unique().to_list())

            sel_modes = st.multiselect("Select Payment Modes", modes, default=[])

            if sel_modes:

                df_display = df_display.filter(pl.col("paymentmode").is_in(sel_modes))

                if hourly_df is not None:

                    hourly_df = hourly_df.filter(pl.col("paymentmode").is_in(sel_modes))



                if any(m in sel_modes for m in CARD_MODES):

                    st.markdown("---")

                    st.subheader("💳 Card Filters")

                    if "card_category" in df_display.columns:

                        cats = sorted(str(x) for x in df_display.get_column("card_category").drop_nulls().unique().to_list())

                        sc   = st.multiselect("Select Card Category (Geo)", cats, default=[])

                        if sc:

                            df_display = df_display.filter(pl.col("card_category").is_in(sc) | ~pl.col("paymentmode").is_in(list(CARD_MODES)))

                            if hourly_df is not None:

                                hourly_df = hourly_df.filter(pl.col("card_category").is_in(sc) | ~pl.col("paymentmode").is_in(list(CARD_MODES)))

                    if "cardtype" in df_display.columns:

                        ctypes = sorted(str(x) for x in df_display.get_column("cardtype").drop_nulls().unique().to_list())

                        sct    = st.multiselect("Select Card Network", ctypes, default=[])

                        if sct:

                            df_display = df_display.filter(pl.col("cardtype").is_in(sct) | ~pl.col("paymentmode").is_in(list(CARD_MODES)))

                            if hourly_df is not None:

                                hourly_df = hourly_df.filter(pl.col("cardtype").is_in(sct) | ~pl.col("paymentmode").is_in(list(CARD_MODES)))



    st.info(f"Ready to analyse **{df_display.height:,}** transactions.")



    if st.button("🚀 Run Report Analysis", key="btn_rep"):

        with st.spinner("Generating Excel files…"):

            num_merchants = df_display.select(pl.col("merchantid").n_unique()).item()



            report_configs = {

                "Overview": {"time_col": None},

                "Daily":    {"time_col": "Day"},

                "Weekly":   {"time_col": "Week"},

                "Monthly":  {"time_col": "Month"},

            }

            if run_hourly and hourly_df is not None and hourly_df.height > 0:

                report_configs["Hourly"] = {"time_col": "Hour", "data": hourly_df}



            base_breakdowns = {

                "Paymode":         ["paymentmode"],

                "PG":              ["pg"],

                "Bank":            ["bankname"],

                "Paymode+PG":      ["paymentmode","pg"],

                "Paymode+Bank":    ["paymentmode","bankname"],

                "Paymode+PG+Bank": ["paymentmode","pg","bankname"],

            }



            def _make_sheet_specs(cur: pl.DataFrame, tcol, report_type):

                tgrp = [tcol] if tcol else []

                specs = {}

                if report_type == "Overview":

                    specs["SR Overall"] = ([], cur)

                else:

                    specs[f"SR {report_type}"] = (tgrp, cur)

                for nm, gcols in base_breakdowns.items():

                    if all(c in cur.columns for c in gcols):

                        specs[f"SR by {nm}"] = (gcols + tgrp, cur)

                specs["SR by Paymode"] = (tgrp + ["paymentmode"], cur)

                if "cardtype" in cur.columns:

                    specs["SR by Card Type"] = (tgrp + ["paymentmode","cardtype"], cur.filter(pl.col("paymentmode").is_in(list(CARD_MODES))))

                if "upi_handle" in cur.columns:

                    hdf = cur.filter(pl.col("upi_handle").is_not_null())

                    if hdf.height > 0:

                        specs["SR by UPI Handle"] = (tgrp+["paymentmode","upi_handle"], hdf)

                if "psp_app" in cur.columns:

                    psp = cur.filter(pl.col("psp_app").is_not_null())

                    if psp.height > 0:

                        specs["SR by PSP App"] = (tgrp+["paymentmode","psp_app"], psp)

                if "bank_tier" in cur.columns:

                    btd = cur.filter(pl.col("paymentmode") == "NET_BANKING")

                    if btd.height > 0:

                        specs["SR by Bank Tier"] = (tgrp+["paymentmode","bank_tier"], btd)

                if "card_category" in cur.columns:

                    ccd = cur.filter(pl.col("paymentmode").is_in(list(CARD_MODES)))

                    if ccd.height > 0:

                        specs["SR by Card Geo"] = (tgrp+["paymentmode","card_category"], ccd)

                if "amount_category" in cur.columns:

                    specs["SR by Amount Category"] = (tgrp+["paymentmode","amount_category"], cur)

                return specs



            def _generate_one(report_type, config):

                cur  = config.get("data", df_display)

                tcol = config["time_col"]

                specs = _make_sheet_specs(cur, tcol, report_type)

                return report_type, _build_report_excel(

                    specs, base_breakdowns, cur, tcol,

                    report_type, num_merchants)



            results = {}

            # Polars processes everything blazing fast. ThreadPool kept strictly for Outer Report Loop

            with ThreadPoolExecutor(max_workers=min(len(report_configs), 4)) as pool:

                futs = {pool.submit(_generate_one, rt, cfg): rt for rt, cfg in report_configs.items()}

                for fut in as_completed(futs):

                    rt, buf = fut.result()

                    results[rt] = buf



            st.markdown("### 📥 Download Reports")

            dl_cols = st.columns(min(len(report_configs), 5))

            for idx, report_type in enumerate(report_configs):

                with dl_cols[idx % len(dl_cols)]:

                    st.download_button(

                        label=f"⬇️ {report_type}",

                        data=results[report_type],

                        file_name=f"{report_type}_SR.xlsx",

                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

                        key=f"dl_{report_type}")

            del results

            gc.collect()

            st.success("✅ All reports ready!")



# ---------------------------------------------------------------------------

# SMALL FILE PANDAS TABS 

# ---------------------------------------------------------------------------

def _decat(df: pd.DataFrame, cols) -> pd.DataFrame:

    df = df.copy(deep=False)

    for c in cols:

        if c in df.columns and hasattr(df[c], "cat"):

            df[c] = df[c].astype(str)

    return df



def compare_periods(curr_df: pd.DataFrame, prev_df: pd.DataFrame, group_cols) -> pd.DataFrame:

    if isinstance(group_cols, str): group_cols = [group_cols]

    curr_df = _decat(curr_df, group_cols)

    prev_df = _decat(prev_df, group_cols)



    sc = (curr_df.groupby(group_cols, observed=True)["is_success"]

          .agg(["count","sum"]).reset_index()

          .rename(columns={"count":"Vol_curr","sum":"Succ_curr"}))

    sp = (prev_df.groupby(group_cols, observed=True)["is_success"]

          .agg(["count","sum"]).reset_index()

          .rename(columns={"count":"Vol_prev","sum":"Succ_prev"}))



    merged = pd.merge(sc, sp, on=group_cols, how="outer")

    num_cols = ["Vol_curr","Succ_curr","Vol_prev","Succ_prev"]

    merged[num_cols] = merged[num_cols].fillna(0)

    merged["SR_curr"]   = merged["Succ_curr"] / merged["Vol_curr"].replace(0,1) * 100

    merged["SR_prev"]   = merged["Succ_prev"] / merged["Vol_prev"].replace(0,1) * 100

    merged["SR_Delta"]  = merged["SR_curr"] - merged["SR_prev"]

    merged["Vol_Delta"] = merged["Vol_curr"] - merged["Vol_prev"]

    return merged.sort_values("SR_Delta")



def get_failure_spike(curr_df: pd.DataFrame, prev_df: pd.DataFrame, group_cols, mode_group: str) -> pd.DataFrame:

    curr_df = _decat(curr_df, group_cols)

    prev_df = _decat(prev_df, group_cols)



    c_err = curr_df[curr_df["txstatus"] != "SUCCESS"].groupby(group_cols, observed=True).size().reset_index(name="curr_count")

    p_err = prev_df[prev_df["txstatus"] != "SUCCESS"].groupby(group_cols, observed=True).size().reset_index(name="prev_count")



    merged = pd.merge(c_err, p_err, on=group_cols, how="outer")

    merged[["curr_count","prev_count"]] = merged[["curr_count","prev_count"]].fillna(0)

    merged["spike"] = merged["curr_count"] - merged["prev_count"]



    total_inc = (len(curr_df[curr_df["txstatus"] != "SUCCESS"]) - len(prev_df[prev_df["txstatus"] != "SUCCESS"]))

    merged["contribution"] = (merged["spike"] / total_inc * 100) if total_inc > 0 else 0.0



    drill = {"UPI": "sub_category", "CARDS": "cardtype", "NET_BANKING": "bankname"}.get(mode_group)

    if drill:

        fails_c = curr_df[curr_df["txstatus"] != "SUCCESS"]

        ctx = []

        for reason in merged["txmsg"]:

            sub = fails_c[fails_c["txmsg"] == reason]

            if not sub.empty and drill in sub.columns:

                top = sub[drill].value_counts().head(3)

                ctx.append(" | ".join(f"{k}: {v}" for k, v in top.items()))

            else:

                ctx.append("N/A")

        merged["Context"] = ctx

    else:

        merged["Context"] = "N/A"



    return merged.sort_values("spike", ascending=False)



def color_delta(val):

    if val < 0:   return "color: #ff4b4b"

    elif val > 0: return "color: #09ab3b"

    return "color: white"



def generate_day_summary(curr_df: pd.DataFrame, prev_df: pd.DataFrame) -> str:

    if curr_df.empty: return "No data."

    cv = len(curr_df); pv = len(prev_df)

    csr  = curr_df["is_success"].sum() / cv * 100 if cv else 0

    psr  = prev_df["is_success"].sum() / pv * 100 if pv else 0

    diff = csr - psr

    if diff >= -0.5:

        return "✅ Performance is stable."

    ud_inc   = int(curr_df["is_userdrop"].sum()) - int(prev_df["is_userdrop"].sum())

    fail_inc = int(curr_df["is_failed"].sum())   - int(prev_df["is_failed"].sum())

    reason   = "User Drops" if ud_inc > fail_inc else "Technical Failures"

    mc = curr_df[curr_df["txstatus"] != "SUCCESS"]["paymentmode"].value_counts()

    top_mode = mc.index[0] if not mc.empty else "Unknown"

    return f"📉 SR dropped by {abs(diff):.1f}% due to increased **{reason}** in **{top_mode}**."



def render_metric_tab(curr_df, prev_df, metric_col, title):

    c_cnt = int(curr_df[metric_col].sum()); p_cnt = int(prev_df[metric_col].sum())

    diff  = c_cnt - p_cnt

    cv = len(curr_df); pv = len(prev_df)

    c_pct = c_cnt / cv * 100 if cv else 0

    p_pct = p_cnt / pv * 100 if pv else 0

    pct_d = c_pct - p_pct



    st.markdown(f'**{title}: {c_pct:.1f}% ({pct_d:+.1f}%) {"📈" if pct_d > 0 else "📉"}**')

    if diff <= 0 and pct_d <= 0:

        st.caption(f"No increase in {title}. (Count: {p_cnt} → {c_cnt})")

        return

    st.markdown("---")



    for lbl, exact, lst, dcol in [

        ("UPI",         "UPI",         None,       "upi_handle"),

        ("CARDS",       None,          CARD_MODES, "cardtype"),

        ("NET_BANKING", "NET_BANKING", None,       "bankname"),

    ]:

        mc = curr_df[curr_df["paymentmode"] == exact] if exact else curr_df[curr_df["paymentmode"].isin(lst)]

        mp = prev_df[prev_df["paymentmode"] == exact] if exact else prev_df[prev_df["paymentmode"].isin(lst)]

        if mc.empty: continue

        mc_cnt = int(mc[metric_col].sum()); mp_cnt = int(mp[metric_col].sum())

        md = mc_cnt - mp_cnt

        if md <= 0: continue

        mc_vol = len(mc); mp_vol = len(mp)

        mc_r = mc_cnt / mc_vol * 100 if mc_vol else 0

        mp_r = mp_cnt / mp_vol * 100 if mp_vol else 0



        c1, c2 = st.columns([1.5, 2])

        with c1:

            st.markdown(f"**{lbl}**")

            st.write(f"Count: {mp_cnt} → {mc_cnt} (+{md})")

            st.caption(f"Rate: {mp_r:.1f}% → {mc_r:.1f}% ({mc_r - mp_r:+.1f}%)")

        with c2:

            if dcol in mc.columns:

                mc2 = _decat(mc[mc[metric_col] == 1], [dcol]); mp2 = _decat(mp[mp[metric_col] == 1], [dcol])

                gc_ = mc2.groupby(dcol, observed=True).size().reset_index(name="Curr")

                gp_ = mp2.groupby(dcol, observed=True).size().reset_index(name="Prev")

                m2  = pd.merge(gc_, gp_, on=dcol, how="outer")

                m2[["Curr","Prev"]] = m2[["Curr","Prev"]].fillna(0)

                m2["Inc"]      = m2["Curr"] - m2["Prev"]

                m2["% Impact"] = (m2["Curr"] / mc_vol * 100).round(2)

                top = m2.sort_values("Inc", ascending=False).head(3)

                if not top.empty and top.iloc[0]["Inc"] > 0:

                    st.dataframe(

                        top.style.format({"Curr":"{:.0f}","Prev":"{:.0f}","Inc":"{:.0f}","% Impact":"{:.2f}%"}),

                        hide_index=True, use_container_width=True)



def render_rca_tab(df: pd.DataFrame):

    st.header("📉 Automated Root Cause Analysis")

    rc1, rc2 = st.columns(2)

    with rc1: sel_days  = st.slider("Comparison Period (Last X Days)", 1, 30, 7)

    with rc2:

        rca_merch = sorted(str(x) for x in df["merchantid"].unique() if pd.notna(x))

        sel_rca_m = st.multiselect("Filter Merchant (RCA)", rca_merch, default=[])



    rca_df  = df[df["merchantid"].astype(str).isin(sel_rca_m)] if sel_rca_m else df

    max_d   = pd.to_datetime(rca_df["Day"].max())

    c_start = max_d   - pd.Timedelta(days=sel_days - 1)

    p_start = c_start - pd.Timedelta(days=sel_days)

    p_end   = c_start - pd.Timedelta(days=1)

    

    st.caption(f"**Current:** {c_start.strftime('%Y-%m-%d')} → {max_d.strftime('%Y-%m-%d')} | **Previous:** {p_start.strftime('%Y-%m-%d')} → {p_end.strftime('%Y-%m-%d')}")



    curr_df = rca_df[(rca_df["Day"] >= c_start) & (rca_df["Day"] <= max_d)]

    prev_df = rca_df[(rca_df["Day"] >= p_start) & (rca_df["Day"] <= p_end)]



    if curr_df.empty or prev_df.empty:

        st.error("Not enough data for the selected range / merchant.")

        return



    cv = len(curr_df); pv = len(prev_df)

    csr  = curr_df["is_success"].sum() / cv * 100 if cv else 0

    psr  = prev_df["is_success"].sum() / pv * 100 if pv else 0

    cgmv = float(curr_df.loc[curr_df["txstatus"]=="SUCCESS","amount"].sum())

    pgmv = float(prev_df.loc[prev_df["txstatus"]=="SUCCESS","amount"].sum())



    m1,m2,m3,m4,m5,m6 = st.columns(6)

    m1.metric("Curr SR",  f"{csr:.2f}%")

    m2.metric("Prev SR",  f"{psr:.2f}%",   f"{csr-psr:.2f}%")

    m3.metric("Curr Vol", f"{cv:,}")

    m4.metric("Prev Vol", f"{pv:,}",        f"{cv-pv:,}")

    m5.metric("Curr GMV", f"₹{cgmv:,.0f}")

    m6.metric("Prev GMV", f"₹{pgmv:,.0f}", f"{cgmv-pgmv:,.0f}")

    st.markdown("---")



    for mode_grp, exact, lst in [("UPI","UPI",None),("CARDS",None,CARD_MODES),("NET_BANKING","NET_BANKING",None)]:

        mc = curr_df[curr_df["paymentmode"]==exact] if exact else curr_df[curr_df["paymentmode"].isin(lst)]

        mp = prev_df[prev_df["paymentmode"]==exact] if exact else prev_df[prev_df["paymentmode"].isin(lst)]



        with st.expander(f"Analysis: {mode_grp}", expanded=(mode_grp=="UPI")):

            if mc.empty:

                st.info("No data for this mode.")

                continue



            sm = compare_periods(mc, mp, "sub_category")

            ws = sm.sort_values("SR_Delta").head(1)



            col_a, col_b = st.columns([1, 1])

            with col_a:

                st.write("**Sub-Category Performance**")

                st.dataframe(

                    sm[["sub_category","SR_curr","SR_prev","SR_Delta","Vol_curr","Vol_prev","Vol_Delta"]]

                    .style.format({"SR_curr":"{:.2f}%","SR_prev":"{:.2f}%","SR_Delta":"{:.2f}%","Vol_curr":"{:.0f}","Vol_prev":"{:.0f}","Vol_Delta":"{:.0f}"})

                    .map(color_delta, subset=["SR_Delta","Vol_Delta"]),

                    use_container_width=True, hide_index=True)



                if not ws.empty and ws["SR_Delta"].values[0] < -1:

                    st.error(f"🚨 **Issue in {ws['sub_category'].values[0]}** (Dropped {ws['SR_Delta'].values[0]:.2f}%)")

                else:

                    st.success("✅ No major sub-category drop.")



                if mode_grp == "UPI" and "upi_handle" in mc.columns:

                    st.markdown("##### 🔍 UPI Handle Breakdown")

                    hs = compare_periods(mc, mp, "upi_handle").sort_values("Vol_Delta")

                    st.dataframe(

                        hs[["upi_handle","SR_curr","SR_prev","SR_Delta","Vol_curr","Vol_prev","Vol_Delta"]]

                        .style.format({"SR_curr":"{:.2f}%","SR_prev":"{:.2f}%","SR_Delta":"{:.2f}%","Vol_curr":"{:.0f}","Vol_prev":"{:.0f}","Vol_Delta":"{:.0f}"})

                        .map(color_delta, subset=["SR_Delta","Vol_Delta"]),

                        use_container_width=True, hide_index=True)



                if mode_grp == "CARDS" and "cardtype" in mc.columns:

                    st.markdown("##### 🔍 Card Type Breakdown")

                    cs = compare_periods(mc, mp, ["paymentmode","cardtype"])

                    st.dataframe(

                        cs[["paymentmode","cardtype","SR_curr","SR_prev","SR_Delta","Vol_curr","Vol_prev","Vol_Delta"]]

                        .style.format({"SR_curr":"{:.2f}%","SR_prev":"{:.2f}%","SR_Delta":"{:.2f}%","Vol_curr":"{:.0f}","Vol_prev":"{:.0f}","Vol_Delta":"{:.0f}"})

                        .map(color_delta, subset=["SR_Delta","Vol_Delta"]),

                        use_container_width=True, hide_index=True)



            with col_b:

                trend = (mc.groupby("Day", observed=True).agg(SR=("is_success","mean"), Vol=("transactionid","count")).reset_index())

                trend["SR"] *= 100

                fig = go.Figure([

                    go.Bar(x=trend["Day"], y=trend["Vol"], name="Volume", marker_color="rgba(135,206,250,0.6)", hovertemplate="<b>Vol:</b> %{y:.0f}<extra></extra>"),

                    go.Scatter(x=trend["Day"], y=trend["SR"], name="SR %", yaxis="y2", line=dict(color="red", width=3), hovertemplate="<b>SR:</b> %{y:.2f}%<extra></extra>"),

                ])

                fig.update_layout(title=f"{mode_grp} Trend", yaxis=dict(title="Volume"), yaxis2=dict(title="SR %", overlaying="y", side="right", range=[0,100]), hovermode="x unified")

                st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})



            st.subheader("Why did it drop?")

            fc1, fc2 = st.columns(2)

            with fc1:

                st.markdown("**1. PG Performance**")

                if "pg" in mc.columns:

                    pg = compare_periods(mc, mp, "pg").sort_values("SR_Delta")

                    st.dataframe(

                        pg[["pg","SR_curr","SR_prev","SR_Delta","Vol_curr","Vol_prev","Vol_Delta"]]

                        .style.format({"SR_curr":"{:.2f}%","SR_prev":"{:.2f}%","SR_Delta":"{:.2f}%","Vol_curr":"{:.0f}","Vol_prev":"{:.0f}","Vol_Delta":"{:.0f}"})

                        .map(color_delta, subset=["SR_Delta","Vol_Delta"]),

                        use_container_width=True, hide_index=True)

            with fc2:

                st.markdown("**2. Error Analysis**")

                spikes = get_failure_spike(mc, mp, ["txmsg"], mode_grp)

                if not spikes.empty:

                    ds = spikes[["txmsg","curr_count","prev_count","spike","contribution","Context"]].copy()

                    ds.columns = ["Error Message","Curr Vol","Prev Vol","Vol Spike","Contrib %","Context"]

                    st.dataframe(

                        ds.style.format({"Curr Vol":"{:.0f}","Prev Vol":"{:.0f}","Vol Spike":"{:.0f}","Contrib %":"{:.2f}%"})

                        .background_gradient(subset=["Vol Spike"], cmap="Reds"),

                        use_container_width=True, hide_index=True)

                else:

                    st.info("No specific error spike detected.")



def render_summary_tab(df: pd.DataFrame):

    st.header("📝 Smart Day-over-Day Analysis")

    ss1, ss2 = st.columns(2)

    

    # Safely get min and max dates

    min_date = df["Day"].min()

    max_date = df["Day"].max()



    with ss1: s_start = st.date_input("Start Date", value=min_date)

    with ss2: s_end   = st.date_input("End Date",   value=max_date)



    # CRITICAL FIX: Convert Streamlit datetime.date outputs to pd.Timestamp to match df["Day"]

    s_start_ts = pd.to_datetime(s_start)

    s_end_ts   = pd.to_datetime(s_end)



    sum_merch = st.multiselect("Filter Merchant (Optional)", sorted(str(x) for x in df["merchantid"].unique() if pd.notna(x)), default=[])

    

    # Filter using correctly typed Timestamps

    sum_df = df[(df["Day"] >= s_start_ts) & (df["Day"] <= s_end_ts)]

    

    if sum_merch: sum_df = sum_df[sum_df["merchantid"].astype(str).isin(sum_merch)]



    if sum_df.empty:

        st.error("No data for selected range.")

        return

    dates = sorted(sum_df["Day"].unique())

    if len(dates) < 2:

        st.warning("Need at least 2 days of data.")

        return



    st.subheader(f"📊 Daily Deep Dives ({len(dates)-1} comparisons)")

    for i in range(1, len(dates)):

        cd = dates[i]; pd_ = dates[i-1]

        dc = sum_df[sum_df["Day"] == cd]

        dp = sum_df[sum_df["Day"] == pd_]



        cv = len(dc); pv = len(dp)

        csr = dc["is_success"].sum()/cv*100 if cv else 0

        psr = dp["is_success"].sum()/pv*100 if pv else 0

        dsr = csr - psr; dv = cv - pv

        icon = "📉" if dsr < -1 else "✅"



        # Safely convert to pd.Timestamp for string formatting

        cd_ts = pd.to_datetime(cd)

        pd_ts = pd.to_datetime(pd_)



        with st.expander(f"{icon} **{cd_ts.strftime('%d-%b')}** (vs {pd_ts.strftime('%d-%b')}) | SR: {csr:.1f}% ({dsr:+.1f}%) | Vol: {pv} → {cv} ({dv:+})"):

            st.info(generate_day_summary(dc, dp))

            t1, t2, t3, t4 = st.tabs(["📉 Success Rate","🚧 User Dropped","💥 Failed","⏳ Incomplete"])

            with t1:

                sm = compare_periods(dc, dp, "paymentmode").sort_values("SR_Delta")

                st.dataframe(

                    sm[["paymentmode","SR_curr","SR_prev","SR_Delta","Vol_curr"]]

                    .style.format({"SR_curr":"{:.1f}%","SR_prev":"{:.1f}%","SR_Delta":"{:+.1f}%","Vol_curr":"{:.0f}"})

                    .map(color_delta, subset=["SR_Delta"]),

                    use_container_width=True, hide_index=True)

            with t2: render_metric_tab(dc, dp, "is_userdrop",   "User Drops")

            with t3: render_metric_tab(dc, dp, "is_failed",     "Technical Failures")

            with t4: render_metric_tab(dc, dp, "is_incomplete", "Incomplete Txns")





# ===========================================================================

# MAIN APP LOOP

# ===========================================================================

def main():

    uploaded_file = st.file_uploader("Upload CSV File", type=["csv"])



    if uploaded_file is not None:

        file_size_mb = uploaded_file.size / (1024 * 1024)

        is_large_file = file_size_mb > LARGE_FILE_THRESHOLD_MB



        if is_large_file:

            st.warning(

                f"📦 Large file detected ({file_size_mb:.1f} MB). "

                f"**Report Generator only** — Insights & RCA and Smart Summary are "

                f"disabled for files above {LARGE_FILE_THRESHOLD_MB} MB to keep performance fast."

            )

        else:

            st.info(f"📄 File size: {file_size_mb:.1f} MB — all features available.")



        # CRITICAL FIX: Use getvalue() instead of read()

        file_bytes = uploaded_file.getvalue()



        with st.spinner("Loading & preprocessing data securely with Polars Engine..."):

            df = load_and_preprocess(file_bytes)

        

        del file_bytes

        gc.collect()



        st.success(

            f"✅ Loaded **{df.height:,}** transactions instantly. "

            f"RAM Used ≈ {df.estimated_size('mb'):.0f} MB"

        )



        # ── LARGE FILE: Report ONLY (pure Polars -> Excel) ─────────────────────

        if is_large_file:

            tab_report, = st.tabs(["📊 Report Generator"])

            with tab_report:

                render_report_tab(df)



        # ── SMALL FILE: All tabs ──────────────────────────────────────────────

        else:

            tab_report, tab_rca, tab_summary = st.tabs([

                "📊 Report Generator", "🔍 Insights & RCA", "📝 Smart Summary"

            ])

            with tab_report:

                render_report_tab(df)



            # Only convert to Pandas at the absolute boundary for the UI tabs 

            # (Memory is safe because we gated this conversion behind is_large_file)

            df_pandas = df.to_pandas()



            # CRITICAL FIX: Ensure the Pandas Date column is consistently cast as a pd.Timestamp Series

            if "Day" in df_pandas.columns:

                df_pandas["Day"] = pd.to_datetime(df_pandas["Day"])

            

            with tab_rca:

                render_rca_tab(df_pandas)

            with tab_summary:

                render_summary_tab(df_pandas)

    else:

        st.info("👆 Upload a CSV file to start.")



if __name__ == "__main__":

    main()
